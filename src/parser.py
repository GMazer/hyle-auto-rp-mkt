# -*- coding: utf-8 -*-
"""
Hyle Auto Report MKT — Excel Parser
Đọc file .xlsx Facebook Ads report, normalize schema, trả về list[dict].
"""

import warnings
from datetime import datetime
from typing import Any

import openpyxl

from src.config import REQUIRED_COLUMNS, KEY_METRICS, logger


def parse_excel_file(filepath: str) -> list[dict[str, Any]]:
    """
    Đọc 1 file Excel báo cáo Facebook Ads.
    Tìm cột bằng tên header (không hardcode index).
    Bỏ qua các cột không nằm trong REQUIRED_COLUMNS.

    Returns:
        list[dict] — mỗi dict là 1 campaign row đã chuẩn hóa.
    """
    # Suppress openpyxl style warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(filepath, data_only=True)

    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError(f"File {filepath} không có sheet nào.")

    # Bước 1: Đọc header row, tìm index theo tên
    header_row = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    col_map: dict[str, int] = {}
    for col_name in REQUIRED_COLUMNS:
        if col_name in header_row:
            col_map[col_name] = header_row.index(col_name)

    # Bước 2: Kiểm tra các cột bắt buộc
    found_cols = set(col_map.keys())
    missing_cols = set(REQUIRED_COLUMNS) - found_cols
    if missing_cols:
        logger.warning(
            "File %s thiếu %d cột: %s",
            filepath, len(missing_cols), missing_cols,
        )

    # Kiểm tra cột spend (bắt buộc)
    spend_col = KEY_METRICS["spend"]
    if spend_col not in col_map:
        wb.close()
        raise ValueError(
            f"File thiếu cột bắt buộc '{spend_col}'. "
            f"Các cột có: {header_row}"
        )

    # Bước 3: Đọc data rows
    records: list[dict[str, Any]] = []
    skipped_rows = 0
    campaign_col = col_map.get("Tên chiến dịch", 0)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if row[campaign_col] is None:
            continue

        record: dict[str, Any] = {"_row": row_idx}
        for col_name, col_idx in col_map.items():
            raw_value = row[col_idx]
            record[col_name] = _clean_value(col_name, raw_value)

        # Validate: ít nhất phải có spend
        if record.get(spend_col) is None or record.get(spend_col) == 0:
            skipped_rows += 1
            logger.debug("Row %d bỏ qua: spend = %s", row_idx, record.get(spend_col))
            continue

        records.append(record)

    wb.close()

    logger.info(
        "Parsed %s: %d campaigns, %d rows bỏ qua, %d cột tìm thấy / %d yêu cầu",
        filepath, len(records), skipped_rows, len(found_cols), len(REQUIRED_COLUMNS),
    )

    return records


def _clean_value(col_name: str, value: Any) -> Any:
    """Chuẩn hóa giá trị theo loại cột."""
    if value is None or value == "" or value == " ":
        return None

    # Cột số
    numeric_cols = {
        "Tần suất",
        "Chi phí trên mỗi kết quả",
        "Số tiền đã chi tiêu (VND)",
        "CPM (Chi phí trên mỗi 1.000 lượt hiển thị)",
        "ROAS kết quả",
        "Chi phí trên mỗi lượt bắt đầu cuộc trò chuyện qua tin nhắn",
        "CTR (Tất cả)",
    }
    if col_name in numeric_cols:
        return _to_float(value)

    # Cột ngày
    date_cols = {"Bắt đầu", "Kết thúc", "Bắt đầu báo cáo", "Kết thúc báo cáo"}
    if col_name in date_cols:
        return _to_date_str(value)

    # Text columns — trả về string
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    """Convert sang float, xử lý string có dấu phẩy."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Xóa dấu phẩy ngăn cách hàng nghìn, thay dấu phẩy thập phân
        cleaned = value.replace(",", "").replace(" ", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _to_date_str(value: Any) -> str | None:
    """Chuẩn hóa date thành string 'YYYY-MM-DD'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value.strip()
    return str(value)


def get_report_date(records: list[dict[str, Any]]) -> str | None:
    """Lấy ngày báo cáo từ records (dùng 'Bắt đầu báo cáo' của row đầu)."""
    if not records:
        return None
    return records[0].get("Bắt đầu báo cáo")
